import numpy as np
# import matplotlib.pyplot as plt
from scipy import linalg
import pandas as pd
from sklearn.linear_model import LogisticRegression
# from sklearn.metrics import confusion_matrix
# import seaborn as sns
import openpyxl


class Process_excel(object):
    def __init__(self, training_data):
        self.training_data = pd.read_excel(training_data)

        self.t = list()
        for row in self.training_data.values:
            self.t.append([row[0]])

        self.inSize = 4
        self.outSize = 1
        self.u = np.zeros((self.inSize, len(self.t)))
        self.y = np.zeros((self.outSize, len(self.t)))
        self.u1 = np.zeros(self.u.shape)
        self.y1 = np.zeros(self.y.shape)

        for i, row in enumerate(self.training_data.values):
            self.u[0, i] = row[4]
            self.u[1, i] = row[5]
            self.u[2, i] = row[6]
            self.u[3, i] = row[7]

            self.y[0, i] = row[12]

    def preprocess(self, delay):
        self.u1 = np.where(self.u < 1000000, self.u, 1000000)
        self.u1 = np.log10(self.u1) - 4.5

        self.y1[0, :delay] = 0
        self.y1[0, delay:] = self.y[0, :self.y.shape[1] - delay]

    # def check(self):
    #     print(self.u[0, 0])
    #     print(self.u[1, 0])
    #     print(self.u[2, 0])
    #     print(self.u[3, 0])
    #     print(self.y[0, 0])
    #     print(self.u1[0, 0])
    #     print(self.y1[0, 0])

    # def out_to_excel(self, save_file):
    #     write_file = openpyxl.Workbook()
    #     sheet = write_file.active
        
    #     sheet.cell(row=1, column=1, value="u_1")
    #     sheet.cell(row=1, column=2, value="u1_1")

    #     for i in range(self.y.shape[1]):
    #         sheet.cell(row=i+2, column=1, value=self.u[0, i])
    #         sheet.cell(row=i+2, column=2, value=self.u1[0, i])
        
    #     write_file.save(save_file)
    #     write_file.close

class ESN(Process_excel):
    def __init__(self, training_data, initLen, resSize, alpha, input_magnitude, spectral_radius):
        super(Process_excel, self).__init__(training_data)
        super(Process_excel, self).preprocess(0)

        self.resSize = resSize
        self.inSize = self.u1.shape[0]
        self.trainLen = self.u1.shape[1]
        self.alpha = alpha
        self.initLen = initLen

        np.random.seed(1)
        self.Win = (np.random.rand(self.resSize, 1 + self.inSize) - 0.5) * input_magnitude
        self.W = np.random.rand(self.resSize, self.resSize) - 0.5 
        rho = max(abs(linalg.eig(self.W)[0]))
        self.W = (self.W / rho) * spectral_radius

        self.X = np.zeros((1 + self.inSize + self.resSize, self.trainLen - initLen))

        self.Yt = np.vstack(self.y1[0, initLen:self.trainLen])
    
    def train(self):
        x = np.zeros((self.resSize, 1))

        for t in range(self.trainLen):
            u_current = self.u1[:, t].reshape((-1, 1))
            x = (1 - self.alpha) * x + self.alpha * np.tanh(np.dot(self.Win, np.vstack((1, u_current))) + np.dot(self.W, x))
            if t >= self.initLen:
                self.X[:, t-self.initLen] = np.vstack((1, u_current, x))[:, 0]

        self.Yt = np.ravel(self.Yt[:, 0:1])
        lr = LogisticRegression(max_iter=10000)
        lr.fit(self.X.T, self.Yt)

        return lr

    def test(self, lr):
        testLen = self.u1.shape[1]
        x = np.zeros((self.resSize, 1))

        for t in range(testLen):
            u_current = self.u1[:, t].reshape((-1, 1))
            x = (1 - self.alpha) * x + self.alpha * np.tanh(np.dot(self.Win, np.vstack((1, u_current))) + np.dot(self.W, x))

            if t >= self.initLen:


training_data = r"/home/a24nitta/work/reservoir/code/learning_data.xlsx"
# save_file = r"/home/a24nitta/work/reservoir/code/check_xlsx"
test = Process_excel(training_data)
test.preprocess(1)
# test.check()
# test.out_to_excel(save_file)
